import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from database import get_db
import models

router = APIRouter()


@router.get("/export/{session_id}")
def export_excel(session_id: int, db: Session = Depends(get_db)):
    session = (
        db.query(models.GenerationSession)
        .filter(models.GenerationSession.id == session_id)
        .first()
    )
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # ── Styles ───────────────────────────────────────────────
    green_fill = PatternFill(start_color="3AAA35", end_color="3AAA35", fill_type="solid")
    white_bold = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    wrap_top = Alignment(wrap_text=True, vertical="top")

    priority_fills = {
        "High":   PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid"),
        "Medium": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
        "Low":    PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    }

    # ── Header row ───────────────────────────────────────────
    headers = [
        "TC ID", "Description", "Type", "Priority",
        "Steps", "Expected Result", "Module", "System", "Feature", "Tester",
    ]
    col_widths = [10, 65, 16, 12, 55, 45, 22, 22, 45, 16]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = green_fill
        cell.font = white_bold
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # ── Data rows ────────────────────────────────────────────
    for row_idx, tc in enumerate(session.test_cases, 2):
        row_data = [
            tc.tc_id,
            tc.description,
            tc.type,
            tc.priority,
            tc.steps or "",
            tc.expected_result or "",
            session.module,
            session.system or "",
            session.feature_title,
            session.tester_name,
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap_top
            if col_idx == 4:
                pf = priority_fills.get(tc.priority)
                if pf:
                    cell.fill = pf

    ws.freeze_panes = "A2"

    # ── Stream response ──────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_module = session.module.replace("/", "-").replace(" ", "_")
    filename = f"Eutelsat_{safe_module}_{session.test_type}_TestCases.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
